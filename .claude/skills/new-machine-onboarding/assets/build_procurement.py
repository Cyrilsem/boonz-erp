#!/usr/bin/env python3
"""
build_procurement.py — Phase 8 of new-machine-onboarding skill.

Reads locked planogram CSVs across all machines for a lead, fans out from
pod_product_name to boonz_product via product_mapping, applies VOX-sourced and
Evian-1L exclusions, applies box-rounding from the procurement memory, and
writes a held procurement_draft.xlsx.

The skill calls this script after Phase 6 lock; CS reviews the workbook;
on approval, lead_deployment_plan.procurement_drafted is set true and the
held PO awaits boonz-master EXECUTE_DEPLOYMENT_PLAN to call create_purchase_order.

USAGE
-----
    python build_procurement.py \
        --lead-slug omd-difc \
        --planogram-dir "BOONZ BRAIN/leads/omd-difc/machines/" \
        --output "BOONZ BRAIN/leads/omd-difc/procurement_draft.xlsx" \
        --supabase-url $SUPABASE_URL \
        --supabase-key $SUPABASE_SERVICE_ROLE_KEY

INPUTS
------
- One planogram.csv per machine in --planogram-dir (subfolders machines/<n>/).
- Each planogram.csv has columns: machine_name, door, shelf, position,
  slot_code, pod_product_name, price, capacity.
- Per-machine config.json sits next to each planogram.csv. Used to read
  venue_group, location_type, source_of_supply.

OUTPUTS
-------
procurement_draft.xlsx with sheets:
  Summary             — totals, supplier count, flag count
  Pod-Level Demand    — per pod_product × machine
  Boonz-Level         — per boonz_product, raw demand and box-rounded
  By Supplier         — one block per supplier, ready for create_purchase_order
  By Brand            — sanity check
  Shelf Capacity      — registry reference
  Validation Flags    — missing mappings, missing avg_cost, missing box size,
                        VOX exclusions applied, Evian-1L exclusions applied
"""

from __future__ import annotations
import argparse
import json
import math
import os
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

# Box-rounding table from procurement_min_order_qty memory.
# When refactoring: source from a database table (e.g. boonz_products.box_size_units)
# and remove this hardcode. Keep in sync with the memory until then.
BOX_SIZES = {
    "Kinder Delice":          10,
    "Kinder Bueno":           10,
    "Oreo Cookies":           12,
    "Snickers":               24,
    "Mars":                   24,
    "M&M Chocolate Nuts":     24,
    "McVities Mini":          12,
    "McVities Dark Chocolate": 12,
    "Nestle KitKat":          24,
    "Nestlé KitKat":          24,
    "Kinder Delice Cake":     10,
    "Bounty":                 24,
    "Barebells":              12,
    "Vitamin Well":           12,
    "Popit":                  24,
    "Krambals":               12,
    "Soft Drinks":             6,   # any soft drink
    "Ice Tea":                 6,   # any ice tea
    "G&H Popped Chips":        8,
    "Perrier":                10,
    "Be Kind Bar":            12,
    "Be Kind Cluster":         8,
    "Tamreem Dates Ball":     25,
    "Evian Regular":          24,
    "Al Ain Regular":         24,
    "Evian 1L":               12,
    "SF Pancake":             10,
    "Red Bull":               24,
    "Zigi":                   14,
}

# VOX-sourced products from reference_vox_sourced_products memory.
# When the schema gains source_of_supply on product_mapping or a sourcing table,
# replace this with a DB query. Keep in sync with memory until then.
VOX_SOURCED_PRODUCTS = {
    "Pepsi Black",
    "Pepsi Regular",
    "Ice Tea Peach",
    "M&M Chocolate Nuts Brown Bag",
    "M&M Chocolate Nuts Yellow Bag",
    "M&M Bag",
    "Aquafina",
    "Maltesers Chocolate Bag",
    "Fade Fit",
    "Chocolate Bar",
    "Skittles Bag",
    "Soft Drinks Mix",
    "7up",
}

EVIAN_1L_NAMES = {"Evian 1L", "Evian - 1L"}
RESTRICTED_LOCATION_TYPES_FOR_EVIAN_1L = {"office", "coworking", "Office", "Co-Working", "Coworking"}

BUFFER_PCT = 0.10  # 10% safety stock above slot capacity sum


# ─── Data shapes ──────────────────────────────────────────────────────────────
@dataclass
class MachineConfig:
    machine_index: int
    planned_official_name: str
    venue_group: Optional[str]
    location_type: Optional[str]
    source_of_supply: str  # 'boonz' | 'partner' | 'mixed'


@dataclass
class PlanogramSlot:
    machine_name: str
    door: str
    shelf: int
    position: int
    slot_code: str
    pod_product_name: str
    price: float
    capacity: int


@dataclass
class ProductMappingRow:
    pod_product_name: str
    boonz_product_id: str
    boonz_product_name: str
    product_brand: Optional[str]
    split_pct: float
    avg_cost: Optional[float]
    is_global_default: bool
    supplier_id: Optional[str]
    supplier_name: Optional[str]
    supplier_code: Optional[str]
    procurement_type: Optional[str]
    box_size: Optional[int] = None  # filled later via lookup


@dataclass
class ValidationFlag:
    severity: str  # 'error' | 'warn' | 'info'
    category: str
    pod_product_name: Optional[str]
    boonz_product_name: Optional[str]
    machine_name: Optional[str]
    msg: str


@dataclass
class ProcurementResult:
    pod_demand: dict = field(default_factory=dict)        # pod_product_name -> total units (post-buffer, pre-fanout)
    boonz_demand_raw: dict = field(default_factory=dict)  # boonz_product_id -> raw units (pre-rounding)
    boonz_demand_rounded: dict = field(default_factory=dict)  # boonz_product_id -> rounded units
    by_supplier: dict = field(default_factory=dict)       # supplier_id -> [boonz_product rows]
    flags: list = field(default_factory=list)
    metadata: dict = field(default_factory=dict)          # boonz_product_id -> ProductMappingRow


# ─── Loaders ──────────────────────────────────────────────────────────────────
def load_planograms(planogram_dir: Path) -> list[tuple[MachineConfig, list[PlanogramSlot]]]:
    """Walk the per-machine subfolders under planogram_dir, load each
    config.json + planogram.csv pair."""
    import csv
    out = []
    for machine_dir in sorted(p for p in planogram_dir.iterdir() if p.is_dir()):
        config_path = machine_dir / "config.json"
        plano_path = machine_dir / "planogram.csv"
        if not config_path.exists() or not plano_path.exists():
            print(f"⚠️  Skipping {machine_dir.name}: missing config.json or planogram.csv", file=sys.stderr)
            continue
        config_data = json.loads(config_path.read_text())
        config = MachineConfig(
            machine_index=int(config_data.get("machine_index", machine_dir.name.split("_")[-1] or 0)),
            planned_official_name=config_data["planned_official_name"],
            venue_group=config_data.get("venue_group"),
            location_type=config_data.get("location_type"),
            source_of_supply=config_data.get("source_of_supply", "boonz"),
        )
        slots = []
        with plano_path.open(newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if not row.get("pod_product_name"):
                    continue  # skip empty slots
                slots.append(PlanogramSlot(
                    machine_name=row["machine_name"],
                    door=row.get("door", "L"),
                    shelf=int(row["shelf"]),
                    position=int(row["position"]),
                    slot_code=row.get("slot_code", ""),
                    pod_product_name=row["pod_product_name"].strip(),
                    price=float(row.get("price") or 0),
                    capacity=int(float(row.get("capacity") or 0)),
                ))
        out.append((config, slots))
    return out


def fetch_product_mappings(supabase_url: str, supabase_key: str, pod_product_names: list[str]) -> list[ProductMappingRow]:
    """Query Supabase REST API for active global-default mappings.
    Skill caller can also dump this from MCP and pass via --mappings-json."""
    import urllib.request
    import urllib.parse

    url = f"{supabase_url}/rest/v1/rpc/get_product_mappings_for_planning"
    body = json.dumps({"p_pod_product_names": pod_product_names}).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=body,
        headers={
            "apikey": supabase_key,
            "Authorization": f"Bearer {supabase_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read())
    except Exception as e:
        print(f"⚠️  Could not fetch mappings via RPC: {e}. Falling back to direct table query.", file=sys.stderr)
        return _fetch_mappings_via_table(supabase_url, supabase_key, pod_product_names)

    return [ProductMappingRow(**r) for r in data]


def _fetch_mappings_via_table(supabase_url: str, supabase_key: str, pod_product_names: list[str]) -> list[ProductMappingRow]:
    """Fallback: hit the product_mapping table directly via PostgREST."""
    import urllib.request
    import urllib.parse

    select = (
        "split_pct,avg_cost,is_global_default,"
        "pod_products(pod_product_name,supplier_id,suppliers(supplier_name,supplier_code,procurement_type)),"
        "boonz_products(product_id,boonz_product_name,product_brand,avg_cost)"
    )
    pod_filter = ",".join(f"\"{n}\"" for n in pod_product_names)
    qs = urllib.parse.urlencode({
        "select": select,
        "is_global_default": "eq.true",
        "status": "eq.Active",
        "pod_products.pod_product_name": f"in.({pod_filter})",
    })
    url = f"{supabase_url}/rest/v1/product_mapping?{qs}"
    req = urllib.request.Request(
        url,
        headers={
            "apikey": supabase_key,
            "Authorization": f"Bearer {supabase_key}",
        },
    )
    with urllib.request.urlopen(req) as resp:
        rows = json.loads(resp.read())

    out = []
    for r in rows:
        pp = r.get("pod_products") or {}
        bp = r.get("boonz_products") or {}
        sup = (pp.get("suppliers") or {})
        out.append(ProductMappingRow(
            pod_product_name=pp.get("pod_product_name", ""),
            boonz_product_id=bp.get("product_id", ""),
            boonz_product_name=bp.get("boonz_product_name", ""),
            product_brand=bp.get("product_brand"),
            split_pct=float(r.get("split_pct") or 100.0),
            avg_cost=r.get("avg_cost") or bp.get("avg_cost"),
            is_global_default=bool(r.get("is_global_default")),
            supplier_id=pp.get("supplier_id"),
            supplier_name=sup.get("supplier_name"),
            supplier_code=sup.get("supplier_code"),
            procurement_type=sup.get("procurement_type"),
        ))
    return out


# ─── Box rounding ─────────────────────────────────────────────────────────────
def lookup_box_size(boonz_product_name: str) -> Optional[int]:
    """Match the boonz_product_name against the BOX_SIZES table.
    Tries exact then prefix match (e.g. 'Vitamin Well Tropical' → 'Vitamin Well')."""
    if not boonz_product_name:
        return None
    if boonz_product_name in BOX_SIZES:
        return BOX_SIZES[boonz_product_name]
    # Prefix match — order longest-first so 'Be Kind Cluster' beats 'Be Kind Bar'
    for key in sorted(BOX_SIZES, key=len, reverse=True):
        if boonz_product_name.startswith(key):
            return BOX_SIZES[key]
    return None


def round_up_to_box(units: float, box_size: Optional[int]) -> int:
    if not box_size or box_size <= 0:
        return math.ceil(units)
    return math.ceil(units / box_size) * box_size


# ─── Core processing ──────────────────────────────────────────────────────────
def build_procurement(
    machines: list[tuple[MachineConfig, list[PlanogramSlot]]],
    mappings: list[ProductMappingRow],
) -> ProcurementResult:
    """Run the full Phase 8 calculation."""
    result = ProcurementResult()

    # Step 1 — pod-level demand per machine, with exclusion checks
    pod_demand_per_machine: dict[tuple[str, str], int] = {}  # (machine_name, pod_product_name) -> capacity
    for config, slots in machines:
        for slot in slots:
            # Evian-1L guardrail
            if slot.pod_product_name in EVIAN_1L_NAMES and config.location_type in RESTRICTED_LOCATION_TYPES_FOR_EVIAN_1L:
                result.flags.append(ValidationFlag(
                    severity="error",
                    category="evian_1l_guardrail",
                    pod_product_name=slot.pod_product_name,
                    boonz_product_name=None,
                    machine_name=config.planned_official_name,
                    msg=f"Evian-1L on {config.location_type} machine — must be removed before lock"
                ))
                continue  # don't include in demand
            key = (config.planned_official_name, slot.pod_product_name)
            pod_demand_per_machine[key] = pod_demand_per_machine.get(key, 0) + slot.capacity

    # Aggregate across machines, apply buffer
    pod_demand_total: dict[str, int] = {}
    for (machine_name, pod_name), capacity in pod_demand_per_machine.items():
        pod_demand_total[pod_name] = pod_demand_total.get(pod_name, 0) + capacity
    for pod_name, units in pod_demand_total.items():
        result.pod_demand[pod_name] = math.ceil(units * (1 + BUFFER_PCT))

    # Step 2 — fan out to boonz level via mappings
    mapping_by_pod: dict[str, list[ProductMappingRow]] = {}
    for m in mappings:
        mapping_by_pod.setdefault(m.pod_product_name, []).append(m)

    # VOX exclusion: build the set of "exclude this boonz_product" decisions per machine
    # If ANY machine in this lead is venue_group=VOX with source_of_supply in (partner,mixed),
    # the VOX-sourced list is excluded from procurement.
    vox_active = any(
        c.venue_group == "VOX" and c.source_of_supply in ("partner", "mixed")
        for c, _ in machines
    )

    for pod_name, demand in result.pod_demand.items():
        rows = mapping_by_pod.get(pod_name, [])
        if not rows:
            result.flags.append(ValidationFlag(
                severity="error",
                category="missing_mapping",
                pod_product_name=pod_name,
                boonz_product_name=None,
                machine_name=None,
                msg=f"No active is_global_default mapping for pod_product '{pod_name}' — procurement total is a lower bound"
            ))
            continue
        for r in rows:
            if vox_active and r.boonz_product_name in VOX_SOURCED_PRODUCTS:
                result.flags.append(ValidationFlag(
                    severity="info",
                    category="vox_exclusion",
                    pod_product_name=pod_name,
                    boonz_product_name=r.boonz_product_name,
                    machine_name=None,
                    msg=f"Excluded from PO — VOX team supplies {r.boonz_product_name}"
                ))
                continue
            allocated = demand * (r.split_pct / 100.0)
            result.boonz_demand_raw[r.boonz_product_id] = result.boonz_demand_raw.get(r.boonz_product_id, 0) + allocated
            result.metadata[r.boonz_product_id] = r
            if r.avg_cost is None:
                result.flags.append(ValidationFlag(
                    severity="warn",
                    category="missing_avg_cost",
                    pod_product_name=pod_name,
                    boonz_product_name=r.boonz_product_name,
                    machine_name=None,
                    msg=f"avg_cost missing for {r.boonz_product_name} — total cost is a lower bound"
                ))

    # Step 3 — box rounding per boonz_product
    for boonz_id, raw in result.boonz_demand_raw.items():
        meta = result.metadata[boonz_id]
        box = lookup_box_size(meta.boonz_product_name)
        meta.box_size = box
        if box is None:
            result.flags.append(ValidationFlag(
                severity="warn",
                category="missing_box_size",
                pod_product_name=None,
                boonz_product_name=meta.boonz_product_name,
                machine_name=None,
                msg=f"No box-size known for {meta.boonz_product_name} — order qty rounded up to nearest unit only; CS to confirm box size"
            ))
        result.boonz_demand_rounded[boonz_id] = round_up_to_box(raw, box)

    # Step 4 — group by supplier
    for boonz_id, qty in result.boonz_demand_rounded.items():
        meta = result.metadata[boonz_id]
        supplier_key = meta.supplier_id or "UNKNOWN"
        result.by_supplier.setdefault(supplier_key, []).append({
            "boonz_product_id": boonz_id,
            "boonz_product_name": meta.boonz_product_name,
            "product_brand": meta.product_brand,
            "supplier_name": meta.supplier_name,
            "supplier_code": meta.supplier_code,
            "procurement_type": meta.procurement_type,
            "qty": qty,
            "raw_qty": result.boonz_demand_raw[boonz_id],
            "box_size": meta.box_size,
            "unit_cost": meta.avg_cost,
            "line_total": (meta.avg_cost or 0) * qty,
        })

    return result


# ─── Excel writer ─────────────────────────────────────────────────────────────
def write_workbook(result: ProcurementResult, output_path: Path, lead_slug: str, machines: list[tuple[MachineConfig, list[PlanogramSlot]]]):
    """Use openpyxl to write the multi-sheet workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4D3A")
    flag_error_fill = PatternFill("solid", fgColor="C1473D")
    flag_warn_fill = PatternFill("solid", fgColor="D97757")
    flag_info_fill = PatternFill("solid", fgColor="C9A25B")
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
            cell.border = border

    def autofit(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    # ── Summary
    ws = wb.active
    ws.title = "Summary"
    total_value = sum(row["line_total"] for rows in result.by_supplier.values() for row in rows)
    total_units = sum(row["qty"] for rows in result.by_supplier.values() for row in rows)
    n_errors = sum(1 for f in result.flags if f.severity == "error")
    n_warns = sum(1 for f in result.flags if f.severity == "warn")
    n_info = sum(1 for f in result.flags if f.severity == "info")

    ws.append(["Boonz · Procurement Draft", "", ""])
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.append(["Lead", lead_slug, ""])
    ws.append(["Machines", len(machines), ""])
    ws.append([])
    ws.append(["Metric", "Value", ""])
    style_header(ws, 5)
    ws.append(["Total units (rounded)", total_units, ""])
    ws.append(["Total cost (AED, lower bound if avg_cost gaps)", round(total_value, 2), ""])
    ws.append(["Distinct boonz_products", len(result.boonz_demand_rounded), ""])
    ws.append(["Distinct suppliers", len(result.by_supplier), ""])
    ws.append([])
    ws.append(["Validation flags", "", ""])
    style_header(ws, ws.max_row)
    ws.append(["Errors", n_errors, "must be resolved before execution" if n_errors else "none"])
    ws.append(["Warnings", n_warns, "review and accept or fix"])
    ws.append(["Info (exclusions)", n_info, "VOX/Evian exclusions applied"])
    autofit(ws)

    # ── Pod-Level Demand
    ws = wb.create_sheet("Pod-Level Demand")
    ws.append(["pod_product_name", "total_demand_post_buffer"])
    style_header(ws)
    for pod_name, units in sorted(result.pod_demand.items(), key=lambda kv: -kv[1]):
        ws.append([pod_name, units])
    autofit(ws)

    # ── Boonz-Level
    ws = wb.create_sheet("Boonz-Level")
    ws.append([
        "boonz_product_id", "boonz_product_name", "brand",
        "supplier_name", "supplier_code", "procurement_type",
        "raw_qty", "box_size", "rounded_qty",
        "unit_cost_aed", "line_total_aed"
    ])
    style_header(ws)
    for boonz_id, qty in sorted(result.boonz_demand_rounded.items(), key=lambda kv: -kv[1]):
        meta = result.metadata[boonz_id]
        unit_cost = meta.avg_cost or 0
        ws.append([
            boonz_id, meta.boonz_product_name, meta.product_brand,
            meta.supplier_name, meta.supplier_code, meta.procurement_type,
            round(result.boonz_demand_raw[boonz_id], 2),
            meta.box_size, qty,
            unit_cost, round(unit_cost * qty, 2)
        ])
    autofit(ws)

    # ── By Supplier
    ws = wb.create_sheet("By Supplier")
    for supplier_id, rows in sorted(result.by_supplier.items(), key=lambda kv: -sum(r["line_total"] for r in kv[1])):
        first = rows[0]
        ws.append([f"Supplier: {first['supplier_name'] or 'UNKNOWN'} ({first['supplier_code'] or '—'}) · {first['procurement_type'] or '—'}"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="C9A25B", size=12)
        ws.append(["boonz_product_id", "boonz_product_name", "qty", "box_size", "unit_cost_aed", "line_total_aed"])
        style_header(ws, ws.max_row)
        subtotal = 0
        for r in sorted(rows, key=lambda x: -x["line_total"]):
            ws.append([
                r["boonz_product_id"], r["boonz_product_name"],
                r["qty"], r["box_size"],
                r["unit_cost"] or 0, round(r["line_total"], 2)
            ])
            subtotal += r["line_total"]
        ws.append(["", "Subtotal", "", "", "", round(subtotal, 2)])
        ws.cell(row=ws.max_row, column=6).font = Font(bold=True)
        ws.append([])  # spacer
    autofit(ws)

    # ── By Brand
    ws = wb.create_sheet("By Brand")
    by_brand: dict[str, dict] = {}
    for boonz_id, qty in result.boonz_demand_rounded.items():
        meta = result.metadata[boonz_id]
        b = meta.product_brand or "(no brand)"
        slot = by_brand.setdefault(b, {"units": 0, "value": 0})
        slot["units"] += qty
        slot["value"] += (meta.avg_cost or 0) * qty
    ws.append(["brand", "total_units", "total_value_aed"])
    style_header(ws)
    for brand, agg in sorted(by_brand.items(), key=lambda kv: -kv[1]["value"]):
        ws.append([brand, agg["units"], round(agg["value"], 2)])
    autofit(ws)

    # ── Shelf Capacity registry (read from assets if available)
    ws = wb.create_sheet("Shelf Capacity")
    registry_path = Path(__file__).parent / "shelf_capacity_registry.csv"
    if registry_path.exists():
        import csv
        with registry_path.open() as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)
        style_header(ws)
        autofit(ws)
    else:
        ws.append(["registry not found — see SKILL.md assets/shelf_capacity_registry.csv"])

    # ── Validation Flags
    ws = wb.create_sheet("Validation Flags")
    ws.append(["severity", "category", "pod_product_name", "boonz_product_name", "machine_name", "message"])
    style_header(ws)
    for f in sorted(result.flags, key=lambda x: ["error", "warn", "info"].index(x.severity)):
        ws.append([f.severity, f.category, f.pod_product_name or "", f.boonz_product_name or "", f.machine_name or "", f.msg])
        fill = flag_error_fill if f.severity == "error" else flag_warn_fill if f.severity == "warn" else flag_info_fill
        for c in ws[ws.max_row]:
            c.fill = fill
            c.font = Font(color="FFFFFF")
    autofit(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅ Wrote {output_path}")
    print(f"   Suppliers: {len(result.by_supplier)} | Units: {total_units} | Value: AED {round(total_value, 2)}")
    print(f"   Flags: {n_errors} errors · {n_warns} warns · {n_info} info")


# ─── CLI ──────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Phase 8 procurement-draft builder")
    ap.add_argument("--lead-slug", required=True)
    ap.add_argument("--planogram-dir", required=True, help="dir containing per-machine subfolders")
    ap.add_argument("--output", required=True, help="path for procurement_draft.xlsx")
    ap.add_argument("--mappings-json", help="optional pre-fetched product_mapping JSON (skip Supabase call)")
    ap.add_argument("--supabase-url", default=os.environ.get("SUPABASE_URL"))
    ap.add_argument("--supabase-key", default=os.environ.get("SUPABASE_SERVICE_ROLE_KEY"))
    args = ap.parse_args()

    plano_dir = Path(args.planogram_dir)
    if not plano_dir.exists():
        print(f"ERROR: {plano_dir} does not exist", file=sys.stderr)
        sys.exit(1)

    machines = load_planograms(plano_dir)
    if not machines:
        print(f"ERROR: no machines found under {plano_dir}", file=sys.stderr)
        sys.exit(1)

    pod_names = sorted({s.pod_product_name for _, slots in machines for s in slots if s.pod_product_name})
    print(f"📋 Loaded {len(machines)} machine(s) covering {len(pod_names)} unique pod_products")

    if args.mappings_json:
        with open(args.mappings_json) as f:
            mapping_rows = [ProductMappingRow(**r) for r in json.load(f)]
    elif args.supabase_url and args.supabase_key:
        mapping_rows = fetch_product_mappings(args.supabase_url, args.supabase_key, pod_names)
    else:
        print("ERROR: provide either --mappings-json or both --supabase-url and --supabase-key", file=sys.stderr)
        sys.exit(1)

    print(f"🔗 Loaded {len(mapping_rows)} product_mapping rows")

    result = build_procurement(machines, mapping_rows)
    write_workbook(result, Path(args.output), args.lead_slug, machines)


if __name__ == "__main__":
    main()
